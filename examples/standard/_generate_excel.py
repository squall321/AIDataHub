"""표준 Excel(.xlsx) 예제 생성 스크립트.

생성되는 sample_data.xlsx 가 시연하는 6원칙
(excel_to_json_conversion_rules.md):

원칙 1 — 시트 상단 고정 (A1 부터 헤더)
원칙 2 — 헤더에 단위 명시 ("무게(g)", "단면적(mm²)" ...)
원칙 3 — 셀 병합 금지
원칙 4 — 색상 의미 별도 컬럼화 (파괴여부 컬럼)
원칙 5 — 1시트 1주제 (Sheet1 만 데이터, _META/_GLOSSARY 는 메타)
원칙 6 — 데이터 의미 명시 (_META + _GLOSSARY 시트)

실행 방법
--------
python _generate_excel.py [출력경로]
기본 출력: 같은 폴더의 sample_data.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook


META_ROWS: list[tuple[str, str]] = [
    ("title", "브라켓 하중 시험 결과 (2026-04)"),
    ("summary", "100개 시료 정적 인장 시험. 표준 KS B 0814."),
    ("tags", "시험,브라켓,하중,2026Q2"),
    ("agents", "material-reviewer,cae-reporter"),
    ("domain", "mechanical-test"),
    ("classification", "internal"),
    ("status", "approved"),
    ("language", "ko"),
    ("source_system", "UTM-500"),
    ("sheet:Sheet1.description", "시료별 최대 하중 측정 결과"),
    ("sheet:Sheet1.method", "KS B 0814 표준 인장시험"),
    ("sheet:Sheet1.condition", "상온 23±2℃, 50±5%RH"),
    ("sheet:Sheet1.equipment", "UTM-500, 50kN 로드셀"),
    ("sheet:Sheet1.operator", "박지수"),
    ("sheet:Sheet1.date", "2026-04-15"),
]


GLOSSARY_ROWS: list[tuple[str, str, str, str]] = [
    ("시료ID", "시료 고유 식별자", "-", "string"),
    ("무게", "시료 무게 (시험 직전 측정)", "g", "float"),
    ("단면적", "시험부 단면적", "mm²", "float"),
    ("최대하중", "시험 중 기록된 최대 하중", "N", "float"),
    ("항복하중", "응력-변형 곡선 항복점 하중", "N", "float"),
    ("파괴여부", "시험 종료 시 파괴 발생 여부", "-", "enum:Y/N"),
]


# 헤더에 단위를 괄호로 명시 (원칙 2)
DATA_HEADERS = [
    "시료ID",
    "무게(g)",
    "단면적(mm²)",
    "최대하중(N)",
    "항복하중(N)",
    "파괴여부",
]

DATA_ROWS = [
    ["S001", 12.34, 25.0, 1250.5, 980.2, "Y"],
    ["S002", 12.28, 25.1, 1242.0, 975.4, "Y"],
    ["S003", 12.41, 24.9, 1268.3, 990.0, "Y"],
    ["S004", 12.30, 25.0, 1255.1, 983.7, "Y"],
    ["S005", 12.36, 25.0, 1260.4, 985.2, "N"],  # 파괴 미발생 (불량)
]


def build_sample_data(out_path: Path) -> Path:
    wb = Workbook()

    # 워크북 빌트인 속성 (12.1 표) — _META 가 있으면 폴백으로만 동작.
    wb.properties.title = "브라켓 하중 시험 결과 (2026-04)"
    wb.properties.subject = "mechanical-test"
    wb.properties.description = (
        "100개 시료 정적 인장 시험. 표준 KS B 0814."
    )
    wb.properties.keywords = "시험,브라켓,하중,2026Q2"
    wb.properties.creator = "CAE팀"
    wb.properties.category = "internal"

    # default sheet → Sheet1 (실제 데이터)
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(DATA_HEADERS)  # 1행 = 헤더 (원칙 1: A1 부터)
    for row in DATA_ROWS:
        ws.append(row)

    # _META 시트 (원칙 6 — 데이터 의미)
    meta = wb.create_sheet("_META")
    meta.append(["key", "value"])
    for k, v in META_ROWS:
        meta.append([k, v])

    # _GLOSSARY 시트 (원칙 6 — 컬럼 의미)
    glos = wb.create_sheet("_GLOSSARY")
    glos.append(["column", "description", "unit", "dtype"])
    for row in GLOSSARY_ROWS:
        glos.append(list(row))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    out = (
        Path(sys.argv[1])
        if len(sys.argv) > 1
        else Path(__file__).resolve().parent / "sample_data.xlsx"
    )
    p = build_sample_data(out)
    print(f"[OK] Excel 예제 생성: {p}  ({p.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
