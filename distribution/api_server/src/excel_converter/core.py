"""Excel(.xlsx) → DATA JSON 변환 핵심 로직.

설계 원칙
- 시트 1개 = DATA JSON 1개 (per_sheet 모드)
- 헤더 1행 추출 → headers 배열
- 본문 → rows (list of list, 각 셀은 number/str/None)
- 병합 셀은 좌상단 값을 모든 셀에 복제 (replicate value)
- 빈 셀은 None 으로 보존
- --infer-units 가 켜지면 headers 와 별도로 units 배열을 만든다
- 빈 시트는 --skip-empty 플래그로 건너뛸 수 있다
- --start-cell A5 처럼 표 시작 셀(좌상단)을 명시할 수 있다 (불규칙 시트용)
- --skip-blank-rows 로 데이터 사이의 빈 행을 제거할 수 있다

caption 은 시트 이름.
data_id 는 ``DATA-{div}-{team}-{year}-{seq:06d}`` 포맷.
"""
from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .units import parse_header_units

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Cell address helpers
# ---------------------------------------------------------------------------

_CELL_ADDR_RE = re.compile(r"^\s*([A-Za-z]+)\s*(\d+)\s*$")


def parse_cell_address(addr: str) -> tuple[int, int]:
    """``"A5"`` 같은 셀 주소를 (row, col) 1-based 튜플로 변환."""
    m = _CELL_ADDR_RE.match(addr)
    if not m:
        raise ValueError(f"invalid cell address: {addr!r}")
    col_letters, row_str = m.group(1), m.group(2)
    col = column_index_from_string(col_letters.upper())
    row = int(row_str)
    if row < 1 or col < 1:
        raise ValueError(f"cell address out of range: {addr!r}")
    return row, col


# ---------------------------------------------------------------------------
# Options / output models
# ---------------------------------------------------------------------------

@dataclass
class XlsxConverterOptions:
    """변환기 옵션."""

    division: str
    team: str
    year: int
    start_seq: int = 1
    output_dir: Path = field(default_factory=lambda: Path("output"))
    mode: str = "per_sheet"            # per_sheet | combined
    skip_empty: bool = False
    skip_blank_rows: bool = False
    infer_units: bool = False
    header_row: int = 1                # 1-based
    start_cell: Optional[str] = None   # 예: "A5" (지정하면 header_row/start_col 무시)
    notes: str = ""                    # 모든 시트에 공통으로 첨부할 메모
    meta_sheet: str = "_META"          # 워크북/시트 컨텍스트가 들어 있는 예약 시트 이름
    glossary_sheet: str = "_GLOSSARY"  # 컬럼 의미 정의 시트 이름

    def __post_init__(self) -> None:
        self.division = self.division.upper()
        self.team = self.team.upper()
        if self.mode not in ("per_sheet", "combined"):
            raise ValueError(f"mode must be 'per_sheet' or 'combined', got {self.mode!r}")
        if self.header_row < 1:
            raise ValueError("header_row must be >= 1")
        if self.start_seq < 0:
            raise ValueError("start_seq must be >= 0")
        if self.start_cell is not None:
            parse_cell_address(self.start_cell)  # 사전 검증
        self.output_dir = Path(self.output_dir)

    def resolve_start(self) -> tuple[int, int]:
        """(start_row, start_col) 을 반환.

        - ``start_cell`` 이 지정되면 그것을 우선 사용.
        - 그렇지 않으면 ``header_row`` 와 column 1 을 사용.
        """
        if self.start_cell:
            return parse_cell_address(self.start_cell)
        return self.header_row, 1


@dataclass
class ConvertedSheet:
    """변환된 시트 1건. write_output() 가 디스크로 떨군다."""

    data_id: str
    caption: str
    headers: list[str]
    rows: list[list[Any]]
    units: Optional[list[Optional[str]]] = None
    notes: str = ""
    source_sheet: str = ""
    warnings: list[str] = field(default_factory=list)
    # ---- 의미 컨텍스트 (원칙 6) ----
    meta_overrides: dict[str, Any] = field(default_factory=dict)
    """워크북 레벨 _META 또는 빌트인 속성에서 추출한 RecordIn 메타. title/summary/tags/agents 등."""
    context: dict[str, Any] = field(default_factory=dict)
    """이 시트에 대한 _META 의 sheet:<name>.<항목>. description/method/condition 등."""
    column_descriptions: dict[str, str] = field(default_factory=dict)
    """_GLOSSARY 의 column → description 매핑 (헤더 매칭된 것만 보존)."""
    units_map: dict[str, str] = field(default_factory=dict)
    """_GLOSSARY 의 column → unit 매핑."""

    def to_payload(self, opts: XlsxConverterOptions) -> dict[str, Any]:
        """JSON 직렬화용 dict 생성."""
        payload: dict[str, Any] = {
            "data_id": self.data_id,
            "schema_version": "data.v1",
            "caption": self.caption,
            "division": opts.division,
            "team": opts.team,
            "year": opts.year,
            "headers": self.headers,
            "rows": self.rows,
            "row_count": len(self.rows),
            "column_count": len(self.headers),
            "source": {
                "sheet": self.source_sheet,
                "kind": "xlsx",
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        if self.units is not None:
            payload["units"] = self.units
        if self.notes:
            payload["notes"] = self.notes
        if self.warnings:
            payload["warnings"] = list(self.warnings)
        # 의미 컨텍스트 (원칙 6) 가 있으면 출력에 반영.
        if self.meta_overrides:
            payload["meta"] = dict(self.meta_overrides)
        if self.context:
            payload["context"] = dict(self.context)
        if self.column_descriptions:
            payload["column_descriptions"] = dict(self.column_descriptions)
        if self.units_map:
            payload["units_map"] = dict(self.units_map)
        return payload


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------

_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+\.\d+$|^-?\.\d+$|^-?\d+\.$|^-?\d+(?:\.\d+)?[eE][+-]?\d+$")


def coerce_value(value: Any) -> Any:
    """openpyxl 셀 값을 JSON 친화적 타입으로 정규화.

    - None → None
    - bool / int / float → 그대로
    - datetime → ISO8601 문자열
    - str → strip 후, 숫자 패턴이면 int/float 으로 변환
    - 기타 → str() 캐스팅
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        if _INT_RE.match(s):
            try:
                return int(s)
            except ValueError:
                return s
        if _FLOAT_RE.match(s):
            try:
                return float(s)
            except ValueError:
                return s
        return s
    return str(value)


# ---------------------------------------------------------------------------
# Merged cell handling
# ---------------------------------------------------------------------------

def _build_merge_lookup(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """병합 영역의 좌상단 값을 (row, col) → value 로 미리 채운다.

    openpyxl 은 병합 영역의 좌상단 셀에만 값을 두고, 나머지 셀은 None 을 반환한다.
    이 함수는 병합 범위 전체에 좌상단 값을 복제한 매핑을 돌려준다.
    """
    lookup: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        anchor = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = anchor
    return lookup


# ---------------------------------------------------------------------------
# Irregular structure detection
# ---------------------------------------------------------------------------

@dataclass
class IrregularReport:
    """시트의 비정상 구조 탐지 결과."""

    is_irregular: bool
    suggested_start_cell: Optional[str]
    reasons: list[str]


def detect_irregular(ws: Worksheet, scan_rows: int = 10, scan_cols: int = 10) -> IrregularReport:
    """첫 ``scan_rows`` × ``scan_cols`` 셀을 훑어 표가 ``A1`` 에서 시작하지 않는 징후를 탐지.

    탐지 규칙 (휴리스틱):

    1. ``A1`` 이 비어 있고, 어딘가에 헤더처럼 보이는 행(모든 셀이 문자열인 비-빈 행)이 있다면
       해당 행의 첫 비-빈 컬럼을 ``suggested_start_cell`` 로 제안.
    2. ``A1`` 에 값이 있어도, 그 행이 모두 문자열이 아니거나 다음 행이 텍스트인 경우
       (제목 셀로 추정) → 더 아래 행을 헤더로 제안.
    3. 모든 후보 셀이 비어 있으면 ``is_irregular=False``, 제안 없음.

    반환:
        IrregularReport(is_irregular, suggested_start_cell, reasons)
    """
    reasons: list[str] = []
    max_row = min(ws.max_row or 0, scan_rows)
    max_col = min(ws.max_column or 0, scan_cols)
    if max_row == 0 or max_col == 0:
        return IrregularReport(False, None, ["sheet is empty"])

    # 셀 값 격자 — 병합 셀도 좌상단 값으로 복제.
    merge_lookup = _build_merge_lookup(ws)
    grid: list[list[Any]] = []
    for r in range(1, max_row + 1):
        row: list[Any] = []
        for c in range(1, max_col + 1):
            if (r, c) in merge_lookup:
                row.append(merge_lookup[(r, c)])
            else:
                row.append(ws.cell(row=r, column=c).value)
        grid.append(row)

    def _is_nonempty(v: Any) -> bool:
        return v is not None and not (isinstance(v, str) and v.strip() == "")

    def _is_str(v: Any) -> bool:
        return isinstance(v, str) and v.strip() != ""

    # 헤더 후보 행 점수: 비-빈 셀이 2개 이상 + 모두 문자열.
    def _looks_like_header(row: list[Any]) -> tuple[bool, int]:
        """(헤더처럼 보이는가, 첫 비-빈 컬럼 인덱스 1-based)"""
        first_col = -1
        nonempty = 0
        all_str = True
        for i, v in enumerate(row, start=1):
            if _is_nonempty(v):
                if first_col == -1:
                    first_col = i
                nonempty += 1
                if not _is_str(v):
                    all_str = False
        return (nonempty >= 2 and all_str, first_col)

    # 1) A1 자체가 비어 있는가?
    a1 = grid[0][0] if grid and grid[0] else None
    a1_empty = not _is_nonempty(a1)

    # 2) A1 에서 시작하는 1행이 헤더 같은가?
    row1_header, row1_first_col = _looks_like_header(grid[0])

    if not a1_empty and row1_header and row1_first_col == 1:
        # 정상: A1 에서 시작하고 1행이 헤더처럼 보임.
        return IrregularReport(False, None, ["A1 starts a header-like row"])

    # 3) 첫 헤더처럼 보이는 행을 찾는다.
    suggested: Optional[str] = None
    for r_idx, row in enumerate(grid, start=1):
        looks, first_col = _looks_like_header(row)
        if looks:
            suggested = f"{get_column_letter(first_col)}{r_idx}"
            reasons.append(
                f"row {r_idx} looks like a header (all-string, {first_col}열부터 시작)"
            )
            break

    if a1_empty:
        reasons.append("A1 is empty")
    elif not row1_header:
        reasons.append("row 1 is not a clean header (not all strings or only 1 cell)")
    elif row1_first_col != 1:
        reasons.append(f"row 1 starts at column {row1_first_col}, not A")

    if suggested is None:
        # 헤더 후보가 없음 — 그래도 비정상 신호는 위에 reasons 로 기록.
        return IrregularReport(bool(reasons), None, reasons)

    is_irregular = suggested != "A1"
    return IrregularReport(is_irregular, suggested, reasons)


# ---------------------------------------------------------------------------
# Meta / Glossary parsers (원칙 6 — 데이터 의미 명시)
# ---------------------------------------------------------------------------

# 원칙 6.1 워크북 레벨 키 — RecordIn 메타로 매핑되는 화이트리스트.
_META_WORKBOOK_KEYS = {
    "title",
    "summary",
    "domain",
    "classification",
    "status",
    "language",
    "source_system",
    "author",
    "department",
    "project",
    "version",
}
# 콤마 분리 → 배열로 변환되는 키.
_META_LIST_KEYS = {"tags", "agents", "subject_keywords"}
# 시트 레벨 키 화이트리스트 (sheet:<name>.<항목>).
_META_SHEET_KEYS = {
    "description",
    "method",
    "condition",
    "equipment",
    "operator",
    "date",
    "notes",
    "caveats",
}


def _split_csv(value: Any) -> list[str]:
    """콤마/세미콜론 분리 + strip + 빈 항목 제거."""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = re.split(r"[,;]", s)
    return [p.strip() for p in parts if p.strip()]


def _parse_meta_sheet(
    ws: Worksheet,
) -> tuple[dict[str, Any], dict[str, dict[str, Any]], list[str]]:
    """`_META` 시트를 파싱.

    구조: 2열 (key | value). 첫 행은 헤더 ``key``, ``value``.
    ``sheet:<sheet_name>.<항목>`` 형식의 키는 시트 레벨 컨텍스트로 분리된다.

    반환:
        (workbook_meta, per_sheet_context, warnings)
        - workbook_meta: ``{title, summary, tags, agents, ...}``
        - per_sheet_context: ``{sheet_name: {description, method, ...}}``
        - warnings: 알 수 없는 키 등 경고 메시지 목록
    """
    workbook_meta: dict[str, Any] = {}
    per_sheet_context: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []

    max_row = ws.max_row or 0
    if max_row == 0:
        return workbook_meta, per_sheet_context, warnings

    # 첫 행이 헤더(key | value) 인 경우 건너뛰기.
    start_row = 1
    first_key = ws.cell(row=1, column=1).value
    if isinstance(first_key, str) and first_key.strip().lower() == "key":
        start_row = 2

    for r in range(start_row, max_row + 1):
        key_raw = ws.cell(row=r, column=1).value
        val_raw = ws.cell(row=r, column=2).value
        if key_raw is None:
            continue
        key = str(key_raw).strip()
        if not key:
            continue
        value = coerce_value(val_raw)
        if value is None or (isinstance(value, str) and not value.strip()):
            # 빈 값은 무시.
            continue

        if key.startswith("sheet:"):
            # sheet:<sheet_name>.<항목> 형식
            tail = key[len("sheet:"):]
            if "." not in tail:
                warnings.append(
                    f"_META: invalid sheet-level key '{key}' (expected 'sheet:<name>.<field>')"
                )
                continue
            sheet_name, field_name = tail.split(".", 1)
            sheet_name = sheet_name.strip()
            field_name = field_name.strip()
            if not sheet_name or not field_name:
                warnings.append(f"_META: invalid sheet-level key '{key}'")
                continue
            if field_name not in _META_SHEET_KEYS:
                warnings.append(
                    f"_META: unknown sheet-level field '{field_name}' "
                    f"(allowed: {sorted(_META_SHEET_KEYS)})"
                )
                # 알 수 없는 필드도 그래도 보존 — 사용자가 정의한 추가 컨텍스트일 수 있음.
            per_sheet_context.setdefault(sheet_name, {})[field_name] = value
        elif key in _META_LIST_KEYS:
            workbook_meta[key] = _split_csv(value)
        elif key in _META_WORKBOOK_KEYS:
            workbook_meta[key] = value
        else:
            warnings.append(
                f"_META: unknown workbook-level key '{key}' (ignored)"
            )

    return workbook_meta, per_sheet_context, warnings


def _parse_glossary_sheet(
    ws: Worksheet,
) -> tuple[dict[str, dict[str, Optional[str]]], list[str]]:
    """`_GLOSSARY` 시트를 파싱.

    구조: 4열 (column | description | unit | dtype). 첫 행은 헤더.

    반환:
        ({column_name: {"description": ..., "unit": ..., "dtype": ...}}, warnings)
    """
    glossary: dict[str, dict[str, Optional[str]]] = {}
    warnings: list[str] = []

    max_row = ws.max_row or 0
    if max_row == 0:
        return glossary, warnings

    # 헤더 검출 — 1행이 ``column`` 으로 시작하면 스킵.
    start_row = 1
    first_cell = ws.cell(row=1, column=1).value
    if isinstance(first_cell, str) and first_cell.strip().lower() == "column":
        start_row = 2

    for r in range(start_row, max_row + 1):
        col_name_raw = ws.cell(row=r, column=1).value
        desc_raw = ws.cell(row=r, column=2).value
        unit_raw = ws.cell(row=r, column=3).value
        dtype_raw = ws.cell(row=r, column=4).value

        if col_name_raw is None:
            continue
        col_name = str(col_name_raw).strip()
        if not col_name:
            continue

        def _clean(v: Any) -> Optional[str]:
            if v is None:
                return None
            s = str(v).strip()
            if not s or s == "-":
                return None
            return s

        glossary[col_name] = {
            "description": _clean(desc_raw),
            "unit": _clean(unit_raw),
            "dtype": _clean(dtype_raw),
        }

    return glossary, warnings


def _extract_workbook_properties(wb: Any) -> dict[str, Any]:
    """openpyxl ``wb.properties`` 에서 RecordIn 메타로 매핑 가능한 빌트인 속성을 추출.

    매핑 (12.1 표 참조):

    - ``title`` → ``meta.title``
    - ``subject`` → ``meta.domain``
    - ``description`` → ``meta.summary``
    - ``category`` → ``meta.classification``
    - ``keywords`` → ``meta.tags`` (콤마 분리)
    - ``creator`` → ``meta.author``
    """
    out: dict[str, Any] = {}
    props = getattr(wb, "properties", None)
    if props is None:
        return out

    def _get(attr: str) -> Optional[str]:
        v = getattr(props, attr, None)
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    title = _get("title")
    if title:
        out["title"] = title
    subject = _get("subject")
    if subject:
        out["domain"] = subject
    description = _get("description")
    if description:
        out["summary"] = description
    category = _get("category")
    if category:
        out["classification"] = category
    keywords = _get("keywords")
    if keywords:
        out["tags"] = _split_csv(keywords)
    creator = _get("creator")
    if creator:
        out["author"] = creator
    return out


def _coerce_with_dtype(value: Any, dtype: Optional[str]) -> tuple[Any, Optional[str]]:
    """``_GLOSSARY`` 의 ``dtype`` 힌트로 셀 값 강제 변환.

    반환: (변환된 값, 경고 메시지 또는 None)
    실패하면 원래 값을 유지하고 경고 메시지를 돌려준다.
    """
    if value is None or dtype is None:
        return value, None
    dtype_lower = dtype.strip().lower()

    try:
        if dtype_lower in ("int", "integer"):
            if isinstance(value, bool):
                return int(value), None
            if isinstance(value, int):
                return value, None
            if isinstance(value, float):
                if value.is_integer():
                    return int(value), None
                return value, f"dtype int 기대했으나 정수가 아닌 실수: {value!r}"
            return int(str(value).strip()), None
        if dtype_lower in ("float", "double", "number", "real"):
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return float(value), None
            return float(str(value).strip()), None
        if dtype_lower in ("bool", "boolean"):
            if isinstance(value, bool):
                return value, None
            s = str(value).strip().lower()
            if s in ("true", "1", "y", "yes"):
                return True, None
            if s in ("false", "0", "n", "no"):
                return False, None
            return value, f"dtype bool 변환 실패: {value!r}"
        if dtype_lower == "string":
            return str(value), None
        if dtype_lower == "date":
            if isinstance(value, datetime):
                return value.isoformat(), None
            return str(value), None
        if dtype_lower.startswith("enum:"):
            allowed = [x.strip() for x in dtype[len("enum:"):].split("/") if x.strip()]
            sval = str(value).strip()
            if sval not in allowed:
                return value, (
                    f"dtype {dtype!r} 위반: {sval!r} 는 {allowed} 에 없음"
                )
            return value, None
    except (ValueError, TypeError) as exc:
        return value, f"dtype {dtype!r} 변환 실패 ({type(exc).__name__}): {value!r}"

    # 알 수 없는 dtype 은 그대로 둔다.
    return value, None


# ---------------------------------------------------------------------------
# Converter
# ---------------------------------------------------------------------------

class XlsxConverter:
    """Excel 워크북을 시트 단위 DATA JSON 으로 변환."""

    def __init__(self, options: XlsxConverterOptions) -> None:
        self.options = options

    # ---- public API --------------------------------------------------

    def convert(self, xlsx_path: Path | str) -> list[ConvertedSheet]:
        """워크북을 읽어 변환된 시트 리스트를 돌려준다.

        원칙 6 (데이터 의미 명시) 처리 흐름:

        1. 빌트인 속성 → 가장 낮은 우선순위 폴백.
        2. ``_META`` 시트 → 워크북 + 시트 레벨 컨텍스트 추출 (빌트인 위에 머지).
        3. ``_GLOSSARY`` 시트 → 컬럼 description / unit / dtype 추출.
        4. 데이터 시트 변환 시 위 정보를 ConvertedSheet 에 반영.
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(xlsx_path)
        if xlsx_path.suffix.lower() != ".xlsx":
            raise ValueError(f"only .xlsx is supported: {xlsx_path}")

        logger.info("loading workbook: %s", xlsx_path)
        wb = load_workbook(xlsx_path, data_only=True, read_only=False)

        # ---- 컨텍스트 추출 (원칙 6) ----------------------------------
        global_warnings: list[str] = []

        # 1) 빌트인 속성 — 폴백.
        wb_properties = _extract_workbook_properties(wb)

        # 2) _META 시트.
        meta_sheet_name = self.options.meta_sheet
        glossary_sheet_name = self.options.glossary_sheet

        workbook_meta: dict[str, Any] = {}
        per_sheet_context: dict[str, dict[str, Any]] = {}
        if meta_sheet_name in wb.sheetnames:
            workbook_meta, per_sheet_context, mw = _parse_meta_sheet(
                wb[meta_sheet_name]
            )
            global_warnings.extend(mw)

        # 머지: 빌트인 < _META (12장 우선순위).
        merged_workbook_meta: dict[str, Any] = dict(wb_properties)
        merged_workbook_meta.update(workbook_meta)

        # 3) _GLOSSARY 시트.
        glossary: dict[str, dict[str, Optional[str]]] = {}
        if glossary_sheet_name in wb.sheetnames:
            glossary, gw = _parse_glossary_sheet(wb[glossary_sheet_name])
            global_warnings.extend(gw)

        # 4) 시트 별 변환 — _META / _GLOSSARY 는 데이터로 처리하지 않는다.
        results: list[ConvertedSheet] = []
        seq = self.options.start_seq
        used_sheet_names: set[str] = set()
        for sheet_name in wb.sheetnames:
            if sheet_name in (meta_sheet_name, glossary_sheet_name):
                logger.info("skipping reserved sheet: %s", sheet_name)
                continue
            ws = wb[sheet_name]
            converted = self._convert_sheet(ws, seq)
            if converted is None:
                logger.info("skipping empty sheet: %s", sheet_name)
                continue

            # 워크북 레벨 메타 머지 — 모든 시트가 동일 워크북 메타를 공유.
            if merged_workbook_meta:
                converted.meta_overrides = dict(merged_workbook_meta)

            # 시트 레벨 컨텍스트 머지.
            sheet_ctx = per_sheet_context.get(sheet_name, {})
            if sheet_ctx:
                converted.context = dict(sheet_ctx)
                used_sheet_names.add(sheet_name)

            # _GLOSSARY 적용 (컬럼 description / unit / dtype).
            if glossary:
                self._apply_glossary(converted, glossary)

            # 워크북 레벨 경고를 첫 시트에만 부여 (combined 모드에서도 보존).
            if global_warnings and not results:
                converted.warnings = list(global_warnings) + list(converted.warnings)

            results.append(converted)
            seq += 1

        # _META 의 sheet:<name> 키 중 실제 시트가 아닌 것 → 경고.
        for sheet_name in per_sheet_context:
            if (
                sheet_name not in used_sheet_names
                and sheet_name not in wb.sheetnames
            ):
                msg = (
                    f"_META: sheet-level keys reference unknown sheet '{sheet_name}' "
                    f"(check spelling)"
                )
                logger.warning(msg)
                if results:
                    results[0].warnings.append(msg)

        if self.options.mode == "combined":
            results = [self._combine(results, self.options.start_seq)] if results else []

        return results

    # ---- glossary application ---------------------------------------

    def _apply_glossary(
        self,
        sheet: ConvertedSheet,
        glossary: dict[str, dict[str, Optional[str]]],
    ) -> None:
        """_GLOSSARY 정의를 ConvertedSheet 에 적용.

        - 헤더 매칭된 컬럼의 description → ``column_descriptions``.
        - unit → ``units_map`` (헤더 인라인 단위보다 우선; 단, 인라인 단위가 있고
          glossary unit 이 없으면 인라인 보존).
        - dtype 힌트 → 모든 행의 해당 컬럼 셀을 강제 변환 시도.
        """
        headers = sheet.headers
        if not headers:
            return
        col_index: dict[str, int] = {h: i for i, h in enumerate(headers)}
        unmatched: list[str] = []

        for col_name, defn in glossary.items():
            if col_name not in col_index:
                unmatched.append(col_name)
                continue
            idx = col_index[col_name]
            description = defn.get("description")
            unit = defn.get("unit")
            dtype = defn.get("dtype")

            if description:
                sheet.column_descriptions[col_name] = description

            if unit:
                # _GLOSSARY 의 unit 가 인라인 단위와 충돌하면 _GLOSSARY 우선 (12장).
                inline_unit: Optional[str] = None
                if sheet.units is not None and idx < len(sheet.units):
                    inline_unit = sheet.units[idx]
                if inline_unit and inline_unit != unit:
                    sheet.warnings.append(
                        f"_GLOSSARY: column '{col_name}' unit '{unit}' overrides "
                        f"inline header unit '{inline_unit}'"
                    )
                sheet.units_map[col_name] = unit
                if sheet.units is not None and idx < len(sheet.units):
                    sheet.units[idx] = unit

            if dtype:
                for r in range(len(sheet.rows)):
                    if idx >= len(sheet.rows[r]):
                        continue
                    new_val, warn = _coerce_with_dtype(sheet.rows[r][idx], dtype)
                    if warn:
                        sheet.warnings.append(
                            f"_GLOSSARY: row {r + 1} column '{col_name}': {warn}"
                        )
                    sheet.rows[r][idx] = new_val

        if unmatched:
            sheet.warnings.append(
                f"_GLOSSARY: columns {unmatched} not found in sheet headers "
                f"{headers}"
            )

    # ---- internals ---------------------------------------------------

    def _data_id(self, seq: int) -> str:
        return f"DATA-{self.options.division}-{self.options.team}-{self.options.year}-{seq:06d}"

    def _convert_sheet(self, ws: Worksheet, seq: int) -> Optional[ConvertedSheet]:
        merge_lookup = _build_merge_lookup(ws)
        warnings: list[str] = []

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            if self.options.skip_empty:
                return None
            return ConvertedSheet(
                data_id=self._data_id(seq),
                caption=ws.title,
                headers=[],
                rows=[],
                units=[] if self.options.infer_units else None,
                notes=self.options.notes,
                source_sheet=ws.title,
                warnings=warnings,
            )

        # 시작 위치 결정.
        header_row, start_col = self.options.resolve_start()

        # 불규칙 구조 자동 탐지 — start_cell 이 지정되지 않은 경우에만.
        if self.options.start_cell is None:
            report = detect_irregular(ws)
            if report.is_irregular and report.suggested_start_cell:
                msg = (
                    f"sheet '{ws.title}' looks irregular (table doesn't start at A1). "
                    f"Suggested --start-cell {report.suggested_start_cell}. "
                    f"Reasons: {'; '.join(report.reasons)}"
                )
                logger.warning(msg)
                warnings.append(msg)

        if header_row > max_row:
            msg = (
                f"sheet '{ws.title}': header_row={header_row} > max_row={max_row}. "
                "Falling back to last row."
            )
            logger.warning(msg)
            warnings.append(msg)
            if self.options.skip_empty:
                return None
            header_row = max_row

        if start_col > max_col:
            msg = (
                f"sheet '{ws.title}': start_col={start_col} > max_col={max_col}. "
                "Falling back to column A."
            )
            logger.warning(msg)
            warnings.append(msg)
            start_col = 1

        # --- headers ---
        headers_raw: list[Any] = []
        for col in range(start_col, max_col + 1):
            v = merge_lookup.get((header_row, col))
            if v is None:
                v = ws.cell(row=header_row, column=col).value
            headers_raw.append(v)

        # 헤더 행에 빈 셀이 섞여 있으면 경고.
        empty_count = sum(
            1 for h in headers_raw if h is None or str(h).strip() == ""
        )
        if headers_raw and empty_count > 0:
            msg = (
                f"sheet '{ws.title}': header row {header_row} has {empty_count}/"
                f"{len(headers_raw)} empty cells. Auto-filling as col_N. "
                "(merged header cells? consider unmerging)"
            )
            logger.warning(msg)
            warnings.append(msg)

        if self.options.infer_units:
            labels: list[str] = []
            units: list[Optional[str]] = []
            for h in headers_raw:
                label, unit = parse_header_units(h)
                if not label:
                    label = f"col_{len(labels) + 1}"
                labels.append(label)
                units.append(unit)
            headers = labels
        else:
            headers = []
            for i, h in enumerate(headers_raw, start=1):
                if h is None or str(h).strip() == "":
                    headers.append(f"col_{i}")
                else:
                    headers.append(str(h).strip())
            units = None

        # --- rows ---
        rows: list[list[Any]] = []
        for row_idx in range(header_row + 1, max_row + 1):
            row_values: list[Any] = []
            any_value = False
            for col in range(start_col, max_col + 1):
                if (row_idx, col) in merge_lookup:
                    raw = merge_lookup[(row_idx, col)]
                else:
                    raw = ws.cell(row=row_idx, column=col).value
                v = coerce_value(raw)
                if v is not None:
                    any_value = True
                row_values.append(v)
            if not any_value:
                # skip_empty (legacy) 또는 skip_blank_rows 옵션이 켜져 있으면 빈 행 제거.
                if self.options.skip_empty or self.options.skip_blank_rows:
                    continue
            rows.append(row_values)

        if self.options.skip_empty and not rows and not any(headers):
            return None

        return ConvertedSheet(
            data_id=self._data_id(seq),
            caption=ws.title,
            headers=headers,
            rows=rows,
            units=units,
            notes=self.options.notes,
            source_sheet=ws.title,
            warnings=warnings,
        )

    def _combine(self, sheets: list[ConvertedSheet], seq: int) -> ConvertedSheet:
        """combined 모드: 시트들을 하나의 ConvertedSheet 로 합친다.

        각 행 앞에 시트 이름 컬럼을 붙여 출처를 보존.
        헤더는 첫 시트 기준이며 다른 시트는 None 으로 패딩한다.
        """
        if not sheets:
            raise ValueError("cannot combine empty list")

        base = sheets[0]
        new_headers = ["__sheet__", *base.headers]
        new_units: Optional[list[Optional[str]]]
        if base.units is not None:
            new_units = [None, *base.units]
        else:
            new_units = None
        new_rows: list[list[Any]] = []
        captions: list[str] = []
        new_warnings: list[str] = []

        width = len(base.headers)
        for s in sheets:
            captions.append(s.caption)
            new_warnings.extend(f"[{s.source_sheet}] {w}" for w in s.warnings)
            for row in s.rows:
                # 폭이 다르면 None 패딩 / 잘라내기.
                if len(row) < width:
                    row = list(row) + [None] * (width - len(row))
                elif len(row) > width:
                    row = list(row[:width])
                new_rows.append([s.source_sheet, *row])

        # 워크북 레벨 메타는 모든 시트가 공유하므로 첫 시트 값을 그대로 가져간다.
        # 시트별 컨텍스트는 source_sheet 별로 모아서 dict-of-dict 로 보존.
        combined_context: dict[str, dict[str, Any]] = {}
        combined_col_descs: dict[str, str] = {}
        combined_units_map: dict[str, str] = {}
        for s in sheets:
            if s.context:
                combined_context[s.source_sheet] = dict(s.context)
            for k, v in s.column_descriptions.items():
                combined_col_descs.setdefault(k, v)
            for k, v in s.units_map.items():
                combined_units_map.setdefault(k, v)

        return ConvertedSheet(
            data_id=self._data_id(seq),
            caption=" | ".join(captions),
            headers=new_headers,
            rows=new_rows,
            units=new_units,
            notes=base.notes,
            source_sheet=",".join(s.source_sheet for s in sheets),
            warnings=new_warnings,
            meta_overrides=dict(base.meta_overrides),
            context=combined_context,
            column_descriptions=combined_col_descs,
            units_map=combined_units_map,
        )


# ---------------------------------------------------------------------------
# Output writing
# ---------------------------------------------------------------------------

def write_output(
    sheet: ConvertedSheet,
    options: XlsxConverterOptions,
) -> Path:
    """ConvertedSheet 를 ``output_dir/data_id.json`` 으로 직렬화."""
    options.output_dir.mkdir(parents=True, exist_ok=True)
    out_path = options.output_dir / f"{sheet.data_id}.json"
    payload = sheet.to_payload(options)
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out_path


def write_outputs(
    sheets: Iterable[ConvertedSheet],
    options: XlsxConverterOptions,
) -> list[Path]:
    """다수 시트를 모두 디스크에 떨군다."""
    return [write_output(s, options) for s in sheets]
