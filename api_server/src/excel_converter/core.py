"""Excel(.xlsx) → DATA JSON 변환 핵심 로직.

설계 원칙
- 시트 1개 = DATA JSON 1개 (per_sheet 모드)
- 헤더 1행 추출 → headers 배열
- 본문 → rows (list of list, 각 셀은 number/str/None)
- 병합 셀은 좌상단 값을 모든 셀에 복제 (replicate value)
- 빈 셀은 None 으로 보존
- --infer-units 가 켜지면 headers 와 별도로 units 배열을 만든다
- 빈 시트는 --skip-empty 플래그로 건너뛸 수 있다

caption 은 시트 이름.
data_id 는 ``DATA-{div}-{team}-{year}-{seq:06d}`` 포맷.
"""
from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .units import parse_header_units

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Options / output models
# ---------------------------------------------------------------------------

@dataclass
class XlsxConverterOptions:
    """변환기 옵션."""

    division: str
    team: str
    year: int
    start_seq: int = 1
    output_dir: Path = field(default_factory=lambda: Path("output"))
    mode: str = "per_sheet"            # per_sheet | combined
    skip_empty: bool = False
    infer_units: bool = False
    header_row: int = 1                # 1-based
    notes: str = ""                    # 모든 시트에 공통으로 첨부할 메모

    def __post_init__(self) -> None:
        self.division = self.division.upper()
        self.team = self.team.upper()
        if self.mode not in ("per_sheet", "combined"):
            raise ValueError(f"mode must be 'per_sheet' or 'combined', got {self.mode!r}")
        if self.header_row < 1:
            raise ValueError("header_row must be >= 1")
        if self.start_seq < 0:
            raise ValueError("start_seq must be >= 0")
        self.output_dir = Path(self.output_dir)


@dataclass
class ConvertedSheet:
    """변환된 시트 1건. write_output() 가 디스크로 떨군다."""

    data_id: str
    caption: str
    headers: list[str]
    rows: list[list[Any]]
    units: Optional[list[Optional[str]]] = None
    notes: str = ""
    source_sheet: str = ""

    def to_payload(self, opts: XlsxConverterOptions) -> dict[str, Any]:
        """JSON 직렬화용 dict 생성."""
        payload: dict[str, Any] = {
            "data_id": self.data_id,
            "schema_version": "data.v1",
            "caption": self.caption,
            "division": opts.division,
            "team": opts.team,
            "year": opts.year,
            "headers": self.headers,
            "rows": self.rows,
            "row_count": len(self.rows),
            "column_count": len(self.headers),
            "source": {
                "sheet": self.source_sheet,
                "kind": "xlsx",
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        if self.units is not None:
            payload["units"] = self.units
        if self.notes:
            payload["notes"] = self.notes
        return payload


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------

_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+\.\d+$|^-?\.\d+$|^-?\d+\.$|^-?\d+(?:\.\d+)?[eE][+-]?\d+$")


def coerce_value(value: Any) -> Any:
    """openpyxl 셀 값을 JSON 친화적 타입으로 정규화.

    - None → None
    - bool / int / float → 그대로
    - datetime → ISO8601 문자열
    - str → strip 후, 숫자 패턴이면 int/float 으로 변환
    - 기타 → str() 캐스팅
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        if _INT_RE.match(s):
            try:
                return int(s)
            except ValueError:
                return s
        if _FLOAT_RE.match(s):
            try:
                return float(s)
            except ValueError:
                return s
        return s
    return str(value)


# ---------------------------------------------------------------------------
# Merged cell handling
# ---------------------------------------------------------------------------

def _build_merge_lookup(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """병합 영역의 좌상단 값을 (row, col) → value 로 미리 채운다.

    openpyxl 은 병합 영역의 좌상단 셀에만 값을 두고, 나머지 셀은 None 을 반환한다.
    이 함수는 병합 범위 전체에 좌상단 값을 복제한 매핑을 돌려준다.
    """
    lookup: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        anchor = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = anchor
    return lookup


# ---------------------------------------------------------------------------
# Converter
# ---------------------------------------------------------------------------

class XlsxConverter:
    """Excel 워크북을 시트 단위 DATA JSON 으로 변환."""

    def __init__(self, options: XlsxConverterOptions) -> None:
        self.options = options

    # ---- public API --------------------------------------------------

    def convert(self, xlsx_path: Path | str) -> list[ConvertedSheet]:
        """워크북을 읽어 변환된 시트 리스트를 돌려준다."""
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(xlsx_path)
        if xlsx_path.suffix.lower() != ".xlsx":
            raise ValueError(f"only .xlsx is supported: {xlsx_path}")

        logger.info("loading workbook: %s", xlsx_path)
        wb = load_workbook(xlsx_path, data_only=True, read_only=False)

        results: list[ConvertedSheet] = []
        seq = self.options.start_seq
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            converted = self._convert_sheet(ws, seq)
            if converted is None:
                logger.info("skipping empty sheet: %s", sheet_name)
                continue
            results.append(converted)
            seq += 1

        if self.options.mode == "combined":
            results = [self._combine(results, self.options.start_seq)] if results else []

        return results

    # ---- internals ---------------------------------------------------

    def _data_id(self, seq: int) -> str:
        return f"DATA-{self.options.division}-{self.options.team}-{self.options.year}-{seq:06d}"

    def _convert_sheet(self, ws: Worksheet, seq: int) -> Optional[ConvertedSheet]:
        merge_lookup = _build_merge_lookup(ws)

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            if self.options.skip_empty:
                return None
            return ConvertedSheet(
                data_id=self._data_id(seq),
                caption=ws.title,
                headers=[],
                rows=[],
                units=[] if self.options.infer_units else None,
                notes=self.options.notes,
                source_sheet=ws.title,
            )

        header_row = self.options.header_row
        if header_row > max_row:
            if self.options.skip_empty:
                return None
            header_row = max_row

        # --- headers ---
        headers_raw: list[Any] = []
        for col in range(1, max_col + 1):
            v = merge_lookup.get((header_row, col))
            if v is None:
                v = ws.cell(row=header_row, column=col).value
            headers_raw.append(v)

        if self.options.infer_units:
            labels: list[str] = []
            units: list[Optional[str]] = []
            for h in headers_raw:
                label, unit = parse_header_units(h)
                if not label:
                    label = f"col_{len(labels) + 1}"
                labels.append(label)
                units.append(unit)
            headers = labels
        else:
            headers = []
            for i, h in enumerate(headers_raw, start=1):
                if h is None or str(h).strip() == "":
                    headers.append(f"col_{i}")
                else:
                    headers.append(str(h).strip())
            units = None

        # --- rows ---
        rows: list[list[Any]] = []
        for row_idx in range(header_row + 1, max_row + 1):
            row_values: list[Any] = []
            any_value = False
            for col in range(1, max_col + 1):
                if (row_idx, col) in merge_lookup:
                    raw = merge_lookup[(row_idx, col)]
                else:
                    raw = ws.cell(row=row_idx, column=col).value
                v = coerce_value(raw)
                if v is not None:
                    any_value = True
                row_values.append(v)
            if not any_value and self.options.skip_empty:
                continue
            rows.append(row_values)

        if self.options.skip_empty and not rows and not any(headers):
            return None

        return ConvertedSheet(
            data_id=self._data_id(seq),
            caption=ws.title,
            headers=headers,
            rows=rows,
            units=units,
            notes=self.options.notes,
            source_sheet=ws.title,
        )

    def _combine(self, sheets: list[ConvertedSheet], seq: int) -> ConvertedSheet:
        """combined 모드: 시트들을 하나의 ConvertedSheet 로 합친다.

        각 행 앞에 시트 이름 컬럼을 붙여 출처를 보존.
        헤더는 첫 시트 기준이며 다른 시트는 None 으로 패딩한다.
        """
        if not sheets:
            raise ValueError("cannot combine empty list")

        base = sheets[0]
        new_headers = ["__sheet__", *base.headers]
        new_units: Optional[list[Optional[str]]]
        if base.units is not None:
            new_units = [None, *base.units]
        else:
            new_units = None
        new_rows: list[list[Any]] = []
        captions: list[str] = []

        width = len(base.headers)
        for s in sheets:
            captions.append(s.caption)
            for row in s.rows:
                # 폭이 다르면 None 패딩 / 잘라내기.
                if len(row) < width:
                    row = list(row) + [None] * (width - len(row))
                elif len(row) > width:
                    row = list(row[:width])
                new_rows.append([s.source_sheet, *row])

        return ConvertedSheet(
            data_id=self._data_id(seq),
            caption=" | ".join(captions),
            headers=new_headers,
            rows=new_rows,
            units=new_units,
            notes=base.notes,
            source_sheet=",".join(s.source_sheet for s in sheets),
        )


# ---------------------------------------------------------------------------
# Output writing
# ---------------------------------------------------------------------------

def write_output(
    sheet: ConvertedSheet,
    options: XlsxConverterOptions,
) -> Path:
    """ConvertedSheet 를 ``output_dir/data_id.json`` 으로 직렬화."""
    options.output_dir.mkdir(parents=True, exist_ok=True)
    out_path = options.output_dir / f"{sheet.data_id}.json"
    payload = sheet.to_payload(options)
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out_path


def write_outputs(
    sheets: Iterable[ConvertedSheet],
    options: XlsxConverterOptions,
) -> list[Path]:
    """다수 시트를 모두 디스크에 떨군다."""
    return [write_output(s, options) for s in sheets]
